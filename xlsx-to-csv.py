import openpyxl
import csv
import sys
import os

def converter_xlsx_para_csv(caminho_xlsx, caminho_csv=None):
    if not os.path.exists(caminho_xlsx):
        print(f"Arquivo não encontrado: {caminho_xlsx}")
        return

    if caminho_csv is None:
        caminho_csv = os.path.splitext(caminho_xlsx)[0] + ".csv"

    print(f"Convertendo: {caminho_xlsx} -> {caminho_csv}")

    # Abre a planilha em modo somente leitura
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True, read_only=True)
    sheet = wb.active

    total_linhas = sheet.max_row  # estimativa para progresso

    with open(caminho_csv, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            writer.writerow([cell if cell is not None else "" for cell in row])

            # Atualiza progresso a cada 500 linhas
            if i % 500 == 0 or i == total_linhas:
                progresso = (i / total_linhas) * 100
                print(f"\rProcessando linha {i}/{total_linhas} ({progresso:.2f}%)", end="")

    print(f"\nConversão finalizada com sucesso! Total de linhas: {i}")

if __name__ == "__main__":
  input = './arquivo.xlsx'
  output = './arquivo.csv'
  
  converter_xlsx_para_csv(input, output)
